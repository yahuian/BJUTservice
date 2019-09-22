'''
1. 函数功能：将excel中的数据写入数据库中
2. 使用方法：将.xlsx类型的excel表格放置在tools目录下,cd ./tools目录下执行python excel-db.py <文件名>
3. 注意事项：excel表格中的数据请提前清洗为如下的格式：
    week    |   time    |   week1   |   week2   |   room
    星期一      第1,2节        1           16        332
'''
import os
import sqlite3
import sys

from openpyxl import load_workbook

# 从命令行读取文件名
if sys.argv[1] == '':
    print('请输入excel表格的文件名')
    exit
else:
    fileName = sys.argv[1]  # sys.argv[0]是该脚本的名称，即excel-db.py

# 从excel表中读数据
try:
    # excel表格的路径及名称，注意只能打开.xlsx类型的
    DstDir = os.getcwd()+"\\"+fileName+'.xlsx'
    wb = load_workbook(DstDir)  # 实例化一个workbook对象，类似于一个excel文件
    sheet = wb.active  # 打开默认的sheet
except FileNotFoundError:
    print('在tools目录下没有找到excel文件')
    raise


row = sheet.max_row        # 行数
column = sheet.max_column  # 列数

tup = ()
templist = []

for r in range(2, row+1):
    l = []
    for c in range(1, column+1):
        value = sheet.cell(r, c).value
        l.append(value)
    tup = (l[0], l[1], l[2], l[3], l[4])
    templist.append(tup)

# 写入数据库
DstDir = os.getcwd()
DATABASE_URI = DstDir+"/freeRoom.db"  # 数据库文件的路径和名称
conn = sqlite3.connect(DATABASE_URI)  # 连接
c = conn.cursor()
# 创建表
c.execute('CREATE TABLE '+fileName +
          ' (week TEXT,time TEXT,week1 INTEGER,week2 INTEGER,room TEXT)')
# 插入数据
c.executemany('INSERT INTO '+fileName +
              ' (week,time,week1,week2,room) VALUES (?,?,?,?,?)', templist)
conn.commit()  # 提交
print('成功写入数据库')
conn.close()  # 关闭
