# 将excel中的数据写入数据库中

import os
import sqlite3
from openpyxl import load_workbook

fileName='classroom4' # excel文档的名字

# 从excel表中读数据
DstDir = os.getcwd()+'/'+fileName+'.xlsx' # excel表格的路径及名称，注意只能打开.xlsx类型的
wb=load_workbook(DstDir) # 实例化一个workbook对象，类似于一个excel文件
sheet=wb.active # 打开默认的sheet

row=sheet.max_row        # 行数
column=sheet.max_column  # 列数

tup=()
templist=[]

for r in range(2,row+1):
    l=[]
    for c in range(1,column+1):
        value=sheet.cell(r,c).value
        l.append(value)
    tup=(l[0],l[1],l[2],l[3],l[4])
    templist.append(tup)

# 写入数据库
DstDir = os.getcwd()
DATABASE_URI = DstDir+"/freeRoom.db" # 数据库文件的路径和名称
conn = sqlite3.connect(DATABASE_URI) # 连接
c = conn.cursor()
# 创建表
c.execute('CREATE TABLE '+fileName+' (week TEXT,time TEXT,begin INTEGER,end INTEGER,room TEXT)')
# 插入数据
c.executemany('INSERT INTO '+fileName+' (week,time,begin,end,room) VALUES (?,?,?,?,?)', templist) 
conn.commit() # 提交
print('成功写入数据库')
conn.close() # 关闭